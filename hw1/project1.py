import cv2
import numpy as np
from PIL import Image
import os
import matplotlib.pyplot as plt
import openpyxl

if not (os.path.exists("./images")):
    os.makedirs('images')

kid_img, fruit_img = cv2.imread(
    "kid blurred-noisy.tif"), cv2.imread("fruit blurred-noisy.tif")

# (a)origin
cv2.imshow('(a)kid_original', kid_img)
kid_img_norm = cv2.normalize(kid_img, None, alpha=0,
                             beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
kid_img_norm = (255*kid_img_norm).astype(np.uint8)
image = Image.fromarray(cv2.cvtColor(kid_img_norm, cv2.COLOR_BGR2RGB))
image.save('./images/(a)kid_original.jpg', dpi=(200.0, 200.0))
# blur
blur = cv2.medianBlur(kid_img, 15)
# blur = cv2.GaussianBlur(kid_img, (9, 9), 0)

# (b)Laplician
Laplician = cv2.Laplacian(blur, cv2.CV_64F, ksize=3)
# take absolute value
abs_Laplacian = cv2.convertScaleAbs(Laplician)
cv2.imshow('(b)kid_Laplacian', abs_Laplacian)
norm_img = cv2.normalize(abs_Laplacian, None, alpha=0,
                         beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
norm_img = (255*norm_img).astype(np.uint8)
image = Image.fromarray(cv2.cvtColor(norm_img, cv2.COLOR_BGR2RGB))
image.save('./images/(b)kid_Laplacian.jpg', dpi=(200.0, 200.0))

# (c)Laplician_sharpened
Laplacian_sharpened = cv2.add(
    kid_img, abs_Laplacian)
cv2.imshow('(c)kid_Laplacian_sharpened', Laplacian_sharpened)
norm_img = cv2.normalize(Laplacian_sharpened, None, alpha=0,
                         beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
norm_img = (255*norm_img).astype(np.uint8)
image = Image.fromarray(cv2.cvtColor(norm_img, cv2.COLOR_BGR2RGB))
image.save('./images/(c)kid_Laplacian_sharpened.jpg', dpi=(200.0, 200.0))

# (d)sobel
sobelx = cv2.Sobel(blur, cv2.CV_64F, 1, 0, ksize=3)
# take absolute value
sobelx = cv2.convertScaleAbs(sobelx)
sobely = cv2.Sobel(blur, cv2.CV_64F, 0, 1, ksize=3)
# take absolute value
sobely = cv2.convertScaleAbs(sobely)
Sobel_gradient = cv2.add(sobelx, sobely)
cv2.imshow('(d)kid_Sobel_gradient', Sobel_gradient)
norm_img = cv2.normalize(Sobel_gradient, None, alpha=0,
                         beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
norm_img = (255*norm_img).astype(np.uint8)
image = Image.fromarray(cv2.cvtColor(norm_img, cv2.COLOR_BGR2RGB))
image.save('./images/(d)kid_Sobel_gradient.jpg', dpi=(200.0, 200.0))

# (e)smoothing_box_filter
smoothing_box_filter = cv2.boxFilter(
    Sobel_gradient, -1, (5, 5), normalize=True)
cv2.imshow('(e)kid_smoothing_box_filter)', smoothing_box_filter)
norm_img = cv2.normalize(smoothing_box_filter, None, alpha=0,
                         beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
norm_img = (255*norm_img).astype(np.uint8)
image = Image.fromarray(cv2.cvtColor(norm_img, cv2.COLOR_BGR2RGB))
image.save('./images/(e)kid_smoothing_box_filter.jpg', dpi=(200.0, 200.0))

f_img = cv2.multiply(smoothing_box_filter, abs_Laplacian)
cv2.imshow('(f)kid_)', f_img)
norm_img = cv2.normalize(f_img, None, alpha=0,
                         beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
norm_img = (255*norm_img).astype(np.uint8)
image = Image.fromarray(cv2.cvtColor(norm_img, cv2.COLOR_BGR2RGB))
image.save('./images/(f)kid.jpg', dpi=(200.0, 200.0))

# (g)high_boost filtering
c = 0.15
g_img = cv2.addWeighted(f_img, c, kid_img, 1-c, 0)
cv2.imshow('(g)kid_)', g_img)
norm_img = cv2.normalize(g_img, None, alpha=0,
                         beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
norm_img = (255*norm_img).astype(np.uint8)
image = Image.fromarray(cv2.cvtColor(norm_img, cv2.COLOR_BGR2RGB))
image.save('./images/(g)kid.jpg', dpi=(200.0, 200.0))

# (h)Power_law
h_img = np.array(255*(g_img/255)**0.85, dtype='uint8')
cv2.imshow('(h)kid_', h_img)
final_norm_img = cv2.normalize(h_img, None, alpha=0,
                               beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
final_norm_img = (255*final_norm_img).astype(np.uint8)
image = Image.fromarray(cv2.cvtColor(norm_img, cv2.COLOR_BGR2RGB))
image.save('./images/(h)kid.jpg', dpi=(200.0, 200.0))

origin_hist = cv2.calcHist([kid_img_norm], [0], None, [256], [0, 256])
final_hist = cv2.calcHist([final_norm_img.astype('uint8')], [
                          0], None, [256], [0, 256])

plt.plot(origin_hist)
plt.title('origin kid_Histogram')
plt.xlabel('graylevel')
plt.ylabel('count')
plt.legend(['origin'])
plt.show()

plt.plot(final_hist)
plt.title('final kid_Histogram')
plt.xlabel('graylevel')
plt.ylabel('count')
plt.legend(['final'])
plt.show()
wb = openpyxl.load_workbook("Histograms.xlsx", data_only=True)
sheet = wb["histograms"]
for i in range(256):
    sheet.cell(i+2, 2).value = int(origin_hist[i])
    sheet.cell(i+2, 3).value = int(final_hist[i])
wb.save("Histograms.xlsx")
# cv2.waitKey()

###################################################################################################################

# (a)origin
cv2.imshow('(a)fruit_original', fruit_img)
fruit_img_norm = cv2.normalize(fruit_img, None, alpha=0,
                               beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
fruit_img_norm = (255*fruit_img_norm).astype(np.uint8)
image = Image.fromarray(cv2.cvtColor(fruit_img_norm, cv2.COLOR_BGR2RGB))
image.save('./images/(a)fruit_original.jpg', dpi=(200.0, 200.0))
# blur
blur = cv2.medianBlur(fruit_img, 15)
# blur = cv2.GaussianBlur(fruit_img, (23, 23), 0)

# (b)Laplician
Laplician = cv2.Laplacian(blur, cv2.CV_64F, ksize=5)
#  take absolute value
abs_Laplacian = cv2.convertScaleAbs(Laplician)
cv2.imshow('(b)fruit_Laplacian', abs_Laplacian)
norm_img = cv2.normalize(abs_Laplacian, None, alpha=0,
                         beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
norm_img = (255*norm_img).astype(np.uint8)
image = Image.fromarray(cv2.cvtColor(norm_img, cv2.COLOR_BGR2RGB))
image.save('./images/(b)fruit_Laplacian.jpg', dpi=(200.0, 200.0))

# (c)Laplician_sharpened
Laplacian_sharpened = cv2.add(
    fruit_img, abs_Laplacian)
cv2.imshow('(c)fruit_Laplacian_sharpened', Laplacian_sharpened)
norm_img = cv2.normalize(Laplacian_sharpened, None, alpha=0,
                         beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
norm_img = (255*norm_img).astype(np.uint8)
image = Image.fromarray(cv2.cvtColor(norm_img, cv2.COLOR_BGR2RGB))
image.save('./images/(c)fruit_Laplacian_sharpened.jpg', dpi=(200.0, 200.0))

# (d)sobel
sobelx = cv2.Sobel(blur, cv2.CV_64F, 1, 0, ksize=3)
# take absolute value
sobelx = cv2.convertScaleAbs(sobelx)
sobely = cv2.Sobel(blur, cv2.CV_64F, 0, 1, ksize=3)
# take absolute value
sobely = cv2.convertScaleAbs(sobely)
Sobel_gradient = cv2.add(sobelx, sobely)
cv2.imshow('(d)fruit_Sobel_gradient', Sobel_gradient)
norm_img = cv2.normalize(Sobel_gradient, None, alpha=0,
                         beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
norm_img = (255*norm_img).astype(np.uint8)
image = Image.fromarray(cv2.cvtColor(norm_img, cv2.COLOR_BGR2RGB))
image.save('./images/(d)fruit_Sobel_gradient.jpg', dpi=(200.0, 200.0))

# (e)smoothing_box_filter
smoothing_box_filter = cv2.boxFilter(
    Sobel_gradient, -1, (5, 5), normalize=True)
cv2.imshow('(e)fruit_smoothing_box_filter)', smoothing_box_filter)
norm_img = cv2.normalize(smoothing_box_filter, None, alpha=0,
                         beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
norm_img = (255*norm_img).astype(np.uint8)
image = Image.fromarray(cv2.cvtColor(norm_img, cv2.COLOR_BGR2RGB))
image.save('./images/(e)fruit_smoothing_box_filter.jpg', dpi=(200.0, 200.0))

f_img = cv2.multiply(smoothing_box_filter, abs_Laplacian)
cv2.imshow('(f)fruit_)', f_img)
norm_img = cv2.normalize(f_img, None, alpha=0,
                         beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
norm_img = (255*norm_img).astype(np.uint8)
image = Image.fromarray(cv2.cvtColor(norm_img, cv2.COLOR_BGR2RGB))
image.save('./images/(f)fruit.jpg', dpi=(200.0, 200.0))

# (g)high_boost filtering
c = 0.15
g_img = cv2.addWeighted(f_img, c, fruit_img, 1-c, 0)
cv2.imshow('(g)fruit_)', g_img)
norm_img = cv2.normalize(g_img, None, alpha=0,
                         beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
norm_img = (255*norm_img).astype(np.uint8)
image = Image.fromarray(cv2.cvtColor(norm_img, cv2.COLOR_BGR2RGB))
image.save('./images/(g)fruit.jpg', dpi=(200.0, 200.0))

# (h)Power_law
h_img = np.array(255*(g_img/255)**0.85, dtype='uint8')
cv2.imshow('(h)fruit_', h_img)
final_norm_img = cv2.normalize(h_img, None, alpha=0,
                               beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
final_norm_img = (255*final_norm_img).astype(np.uint8)
image = Image.fromarray(cv2.cvtColor(norm_img, cv2.COLOR_BGR2RGB))
image.save('./images/(h)fruit.jpg', dpi=(200.0, 200.0))

origin_hist = cv2.calcHist([fruit_img_norm], [0], None, [256], [0, 256])
final_hist = cv2.calcHist([final_norm_img.astype('uint8')], [
                          0], None, [256], [0, 256])
plt.plot(origin_hist)
plt.title('origin fruit_Histogram')
plt.xlabel('graylevel')
plt.ylabel('count')
plt.legend(['origin'])
plt.show()

plt.plot(final_hist)
plt.title('final fruit_Histogram')
plt.xlabel('graylevel')
plt.ylabel('count')
plt.legend(['final'])
plt.show()

wb = openpyxl.load_workbook("Histograms.xlsx", data_only=True)
sheet = wb["histograms"]
for i in range(256):
    sheet.cell(i+2, 4).value = int(origin_hist[i])
    sheet.cell(i+2, 5).value = int(final_hist[i])
wb.save("Histograms.xlsx")

cv2.waitKey()
